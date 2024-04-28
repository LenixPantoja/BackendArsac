from django.shortcuts import get_object_or_404, render
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework import status
from django.http import Http404
from django.contrib.auth.hashers import make_password
from django.contrib.auth.models import User
from rest_framework import generics
from rest_framework import viewsets
from django.http import JsonResponse
from rest_framework import permissions, status
from rest_framework.decorators import api_view, permission_classes

from dateutil.parser import isoparse
from datetime import datetime

from AppAsistencia.models import *
from AppUsuarios.models import *
from AppAsistencia.serializers import *
from AppReportes.serializers import *

# Librerias para generar pdf o excel
from io import BytesIO
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from django.http import HttpResponse
from openpyxl import Workbook
from reportlab.pdfgen import canvas

# Create your views here.

class NotificationListAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    def get(self, request, format=None):
        notifications = Notification.objects.all()
        serializer = NotificationSerializer(notifications, many=True)
        return Response(serializer.data)

    def post(self, request, format=None):
        serializer = NotificationSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({"Msj":"Notificacion creada exitosamente."}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class NotificationDetailAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    def get_object(self, pk):
        try:
            return Notification.objects.get(pk=pk)
        except Notification.DoesNotExist:
            raise Http404

    def get(self, request, pk, format=None):
        notification = self.get_object(pk)
        serializer = NotificationSerializer(notification)
        return Response(serializer.data)

    def put(self, request, pk, format=None):
        notification = self.get_object(pk)
        serializer = NotificationSerializer(notification, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({"Msj":"Se actualizó la notificación exitosamente."})
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk, format=None):
        notification = self.get_object(pk)
        notification.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    
class AppApiReportesPorEstudiante(APIView):
    def get(self, request, format=None):
        asistencia =  AsistenciaEstudiante.objects.all()
        observaciones = ObservacionesEstudiante.objects.all()
        serializer = AsistenciaEstudianteSerializer(asistencia, many=True)
        #Parametros para obtener el reporte filtrado.
        pUser = request.query_params.get("pUser", None)
        pNumeroDocumento = request.query_params.get("pNumeroDocumento", None)
        pMateria =  int(request.query_params.get("pMateria", None))
        pCurso = int(request.query_params.get("pCurso", None))
        pRango1 = request.query_params.get("pRango1", None)
        pRango2 = request.query_params.get("pRango2", None)
        

        dataReporte = []
        # Parametro de consulta
        if pNumeroDocumento is None or pMateria is None or pRango1 is None or pRango2 is None or pCurso is None:
            return Response(
                {"error": "Parametros incompletos(pNumeroDocumento, pMateria, pRango1, pRango2)"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        # Recorridos
        for dataAsistencia in serializer.data:
            matricula = Matricula.objects.get(id=dataAsistencia["matricula_estudiante"])
            estudiante = Estudiante.objects.get(id=matricula.estudiante.id)
            cedulaEst =  estudiante.estudiante_numero_Id
            curso = Curso.objects.get(id=matricula.curso_Materia.curso.id)
            idCurso = curso.id
            idMateria = matricula.curso_Materia.materia.id
            periodo = Periodo.objects.get(id=matricula.curso_Materia.materia.periodo.id)
            fecha_objeto = datetime.strptime(dataAsistencia["asistenciaEst_created_at"], '%Y-%m-%dT%H:%M:%S.%fZ')
            fecha_formateada = fecha_objeto.strftime('%Y-%m-%d')
            usuarioDocente = matricula.curso_Materia.materia.docente.user.username

            if (fecha_formateada >= pRango1 and 
                fecha_formateada <= pRango2 and 
                pNumeroDocumento == cedulaEst and 
                pMateria == idMateria and 
                idCurso == pCurso and 
                usuarioDocente == pUser):

                observaciones = ObservacionesEstudiante.objects.all()
                lista_observaciones = []
                contador = 0
                for observ in observaciones:
                    if observ.asistenciaEst.id == dataAsistencia["id"]:
                        contador=contador + 1
                        lista_observaciones.append(f"Observación {contador}:{observ.observacionEst}")
                #Formateo de fechas
                hora_inicio = matricula.curso_Materia.materia.horario.hora_inicio
                hora_inicio_format = hora_inicio.strftime("%I:%M:%S %p")
                hora_fin = matricula.curso_Materia.materia.horario.hora_fin
                hora_fin_format = hora_fin.strftime("%I:%M:%S %p")
                
                dataReporte.append({
                    "id": dataAsistencia["id"],
                    "Tipo_asistencia": dataAsistencia["tipo_asistencia"],
                    "Descripcion": dataAsistencia["descripcion"],
                    "Hora_llegada": dataAsistencia["hora_llegada"],
                    "FechaCreacion": dataAsistencia["asistenciaEst_created_at"],
                    "Nombre_estudiante":estudiante.user.first_name + " " + estudiante.user.last_name,
                    "Curso_matriculado": curso.nombre_curso,
                    "Periodo": periodo.nombre_periodo,
                    "Materia": matricula.curso_Materia.materia.nombre_materia,
                    "Horario": f"{hora_inicio_format} a {hora_fin_format}",
                    "Docente": matricula.curso_Materia.materia.docente.user.first_name + " " + matricula.curso_Materia.materia.docente.user.last_name,
                    "Lista_Observaciones": lista_observaciones
                })
        return Response(dataReporte)
    
class AppApiReportesPorEstudiantePDF(APIView):

    def get(self, request, format=None):
        asistencia =  AsistenciaEstudiante.objects.all()
        observaciones = ObservacionesEstudiante.objects.all()
        serializer = AsistenciaEstudianteSerializer(asistencia, many=True)
        #Parametros para obtener el reporte filtrado.
        pUser =  request.query_params.get("pUser", None)
        pNumeroDocumento = request.query_params.get("pNumeroDocumento", None)
        pMateria =  int(request.query_params.get("pMateria", None))
        pCurso = int(request.query_params.get("pCurso", None))
        pRango1 = request.query_params.get("pRango1", None)
        pRango2 = request.query_params.get("pRango2", None)
        
        dataReporte = []
        # Parametro de consulta
        if pNumeroDocumento is None or pMateria is None or pRango1 is None or pRango2 is None or pCurso is None:
            return Response(
                {"error": "Parametros incompletos(pNumeroDocumento, pMateria, pRango1, pRango2, pUser)"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        # Recorridos
        for dataAsistencia in serializer.data:
            matricula = Matricula.objects.get(id=dataAsistencia["matricula_estudiante"])
            estudiante = Estudiante.objects.get(id=matricula.estudiante.id)
            cedulaEst =  estudiante.estudiante_numero_Id
            curso = Curso.objects.get(id=matricula.curso_Materia.curso.id)
            idCurso = curso.id
            idMateria = matricula.curso_Materia.materia.id
            periodo = Periodo.objects.get(id=matricula.curso_Materia.materia.periodo.id)
            fecha_objeto = datetime.strptime(dataAsistencia["asistenciaEst_created_at"], '%Y-%m-%dT%H:%M:%S.%fZ')
            fecha_formateada = fecha_objeto.strftime('%Y-%m-%d')
            usuarioDocente = matricula.curso_Materia.materia.docente.user.username
            if (fecha_formateada >= pRango1 and 
                fecha_formateada <= pRango2 and 
                pNumeroDocumento == cedulaEst and 
                pMateria == idMateria and 
                idCurso == pCurso and
                usuarioDocente == pUser):
                observaciones = ObservacionesEstudiante.objects.all()
                lista_observaciones = []
                contador = 0
                for observ in observaciones:
                    if observ.asistenciaEst.id == dataAsistencia["id"]:
                        contador=contador + 1
                        lista_observaciones.append(f"Observación {contador}:{observ.observacionEst}")
                #Formateo de fechas
                hora_inicio = matricula.curso_Materia.materia.horario.hora_inicio
                hora_inicio_format = hora_inicio.strftime("%I:%M:%S %p")
                hora_fin = matricula.curso_Materia.materia.horario.hora_fin
                hora_fin_format = hora_fin.strftime("%I:%M:%S %p")
                
                dataReporte.append({
                    "id": dataAsistencia["id"],
                    "Tipo_asistencia": dataAsistencia["tipo_asistencia"],
                    "Descripcion": dataAsistencia["descripcion"],
                    "Hora_llegada": dataAsistencia["hora_llegada"],
                    "FechaCreacion": dataAsistencia["asistenciaEst_created_at"],
                    "Nombre_estudiante":estudiante.user.first_name + " " + estudiante.user.last_name,
                    "Cc_estudiante": estudiante.estudiante_numero_Id,
                    "Curso_matriculado": curso.nombre_curso,
                    "Periodo": periodo.nombre_periodo,
                    "Materia": matricula.curso_Materia.materia.nombre_materia,
                    "Horario": f"{hora_inicio_format} a {hora_fin_format}",
                    "Docente": matricula.curso_Materia.materia.docente.user.first_name + " " + matricula.curso_Materia.materia.docente.user.last_name,
                    "Lista_Observaciones": lista_observaciones
                })
        # Creamos un objeto BytesIO para almacenar el PDF generado
        buffer = BytesIO()
        # Creamos un objeto SimpleDocTemplate con el buffer como archivo de destino y tamaño de página oficio
        pdf = SimpleDocTemplate(buffer, pagesize=landscape(letter))

        # Lista de datos
        tabla_datos = []
        # Encabezado de la tabla
        encabezado = ["Tipo de asistencia", "Descripción", "Hora de llegada",
                       "Nombre del estudiante", "Cédula del estudiante", "Curso matriculado",
                       "Materia", "Horario", "Docente", "Lista de Observaciones"]

        tabla_datos.append(encabezado)
        for item in dataReporte:
            fila = [
                #item["id"],
                item["Tipo_asistencia"],
                item["Descripcion"],
                item["Hora_llegada"],
                #item["FechaCreacion"],
                item["Nombre_estudiante"],
                item["Cc_estudiante"],
                item["Curso_matriculado"],
                #item["Periodo"],
                item["Materia"],
                item["Horario"],
                item["Docente"],
                '\n'.join(item["Lista_Observaciones"])  # Convertimos la lista de observaciones a una cadena separada por saltos de línea
        ]
            tabla_datos.append(fila)
        tabla = Table(tabla_datos)

        # Aplicamos estilos a la tabla
        estilo = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),  # Color de fondo para el encabezado
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),  # Color de texto para el encabezado
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Alineación centrada para todas las celdas
            ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),  # Grosor y color de la línea interior de las celdas
            ('BOX', (0, 0), (-1, -1), 0.25, colors.black),  # Grosor y color del borde de las celdas
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Alineación vertical centrada
            ('WORDWRAP', (0, 0), (-1, -1), True),  # Ajuste de texto automático
        ])

        tabla.setStyle(estilo)

        # Agregamos la tabla al documento PDF
        elementos = []
        elementos.append(tabla)

        # Construimos el documento PDF
        pdf.build(elementos)

        # Volvemos al principio del buffer
        buffer.seek(0)

        # Creamos una respuesta HTTP con el contenido del buffer como archivo adjunto
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="archivo.pdf"'

        # Escribimos el contenido del buffer en la respuesta HTTP
        response.write(buffer.getvalue())

        return response


class AppApiReportesPorEstudianteXLSX(APIView):

    def get(self, request, format=None):
        asistencia =  AsistenciaEstudiante.objects.all()
        observaciones = ObservacionesEstudiante.objects.all()
        serializer = AsistenciaEstudianteSerializer(asistencia, many=True)
        #Parametros para obtener el reporte filtrado.
        pUser =  request.query_params.get("pUser", None)
        pNumeroDocumento = request.query_params.get("pNumeroDocumento", None)
        pMateria =  int(request.query_params.get("pMateria", None))
        pCurso = int(request.query_params.get("pCurso", None))
        pRango1 = request.query_params.get("pRango1", None)
        pRango2 = request.query_params.get("pRango2", None)
        
        dataReporte = []
        # Parametro de consulta
        if pNumeroDocumento is None or pMateria is None or pRango1 is None or pRango2 is None or pCurso is None:
            return Response(
                {"error": "Parametros incompletos(pNumeroDocumento, pMateria, pRango1, pRango2, pUser)"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        # Recorridos
        for dataAsistencia in serializer.data:
            matricula = Matricula.objects.get(id=dataAsistencia["matricula_estudiante"])
            estudiante = Estudiante.objects.get(id=matricula.estudiante.id)
            cedulaEst =  estudiante.estudiante_numero_Id
            curso = Curso.objects.get(id=matricula.curso_Materia.curso.id)
            idCurso = curso.id
            idMateria = matricula.curso_Materia.materia.id
            periodo = Periodo.objects.get(id=matricula.curso_Materia.materia.periodo.id)
            fecha_objeto = datetime.strptime(dataAsistencia["asistenciaEst_created_at"], '%Y-%m-%dT%H:%M:%S.%fZ')
            fecha_formateada = fecha_objeto.strftime('%Y-%m-%d')
            usuarioDocente = matricula.curso_Materia.materia.docente.user.username
            if (fecha_formateada >= pRango1 and 
                fecha_formateada <= pRango2 and 
                pNumeroDocumento == cedulaEst and 
                pMateria == idMateria and 
                idCurso == pCurso and
                usuarioDocente == pUser):
                observaciones = ObservacionesEstudiante.objects.all()
                lista_observaciones = []
                contador = 0
                for observ in observaciones:
                    if observ.asistenciaEst.id == dataAsistencia["id"]:
                        contador=contador + 1
                        lista_observaciones.append(f"Observación {contador}:{observ.observacionEst}")
                #Formateo de fechas
                hora_inicio = matricula.curso_Materia.materia.horario.hora_inicio
                hora_inicio_format = hora_inicio.strftime("%I:%M:%S %p")
                hora_fin = matricula.curso_Materia.materia.horario.hora_fin
                hora_fin_format = hora_fin.strftime("%I:%M:%S %p")
                
                dataReporte.append({
                    "id": dataAsistencia["id"],
                    "Tipo_asistencia": dataAsistencia["tipo_asistencia"],
                    "Descripcion": dataAsistencia["descripcion"],
                    "Hora_llegada": dataAsistencia["hora_llegada"],
                    "FechaCreacion": dataAsistencia["asistenciaEst_created_at"],
                    "Nombre_estudiante":estudiante.user.first_name + " " + estudiante.user.last_name,
                    "Cc_estudiante": estudiante.estudiante_numero_Id,
                    "Curso_matriculado": curso.nombre_curso,
                    "Periodo": periodo.nombre_periodo,
                    "Materia": matricula.curso_Materia.materia.nombre_materia,
                    "Horario": f"{hora_inicio_format} a {hora_fin_format}",
                    "Docente": matricula.curso_Materia.materia.docente.user.first_name + " " + matricula.curso_Materia.materia.docente.user.last_name,
                    "Lista_Observaciones": lista_observaciones
                })
        # Creamos un libro de trabajo de Excel
        wb = Workbook()
        ws = wb.active
        # Creamos el encabezado
        header = ["ID", "Tipo de asistencia", "Descripción", "Hora de llegada", "Fecha de creación",
                    "Nombre del estudiante", "Cédula del estudiante", "Curso matriculado", "Período",
                    "Materia", "Horario", "Docente"]

        # Agregamos el encabezado a la primera fila
        ws.append(header)
        # Añadimos los datos a las filas
        for item in dataReporte:
            # Creamos una fila para cada registro
            row = [
                item["id"],
                item["Tipo_asistencia"],
                item["Descripcion"],
                item["Hora_llegada"],
                item["FechaCreacion"],
                item["Nombre_estudiante"],
                item["Cc_estudiante"],
                item["Curso_matriculado"],
                item["Periodo"],
                item["Materia"],
                item["Horario"],
                item["Docente"]
            ]
            
            # Agregamos las observaciones en columnas adicionales
            for i, observacion in enumerate(item["Lista_Observaciones"]):
                row.append(observacion)

            # Añadimos la fila al libro de trabajo
            ws.append(row)

        # Creamos la respuesta HTTP
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="archivo.xlsx"'

        # Guardamos el libro de trabajo en la respuesta
        wb.save(response)

        return response



class AppApiReporteDiario(APIView):
    
    def get(self, request, format=None):
        asistencia = AsistenciaEstudiante.objects.all()
        observaciones = ObservacionesEstudiante.objects.all()
        serializer = AsistenciaEstudianteSerializer(asistencia, many=True)
        pUser = request.query_params.get("pUser", None)
        pRango1 = request.query_params.get("pRango1", None)
        pRango2 = request.query_params.get("pRango2", None)
        dataReporte = []

        if pRango1 is None or pRango2 is None or pUser is None:
            return Response(
                {"error": "Parametros incompletos(pRango1, pRango2, pUser)"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        
        for dataAsistencia in serializer.data:
            
            matricula = Matricula.objects.get(id=dataAsistencia["matricula_estudiante"])
            estudiante = Estudiante.objects.get(id=matricula.estudiante.id)
            cedulaEst = estudiante.estudiante_numero_Id
            curso = Curso.objects.get(id=matricula.curso_Materia.curso.id)
            periodo = Periodo.objects.get(id=matricula.curso_Materia.materia.periodo.id)
            fecha_objeto = datetime.strptime(dataAsistencia["asistenciaEst_created_at"], '%Y-%m-%dT%H:%M:%S.%fZ')
            observaciones_asistencia = observaciones.filter(asistenciaEst_id=dataAsistencia["id"])  # Filtrar observaciones por asistencia
            fecha_formateada = fecha_objeto.strftime('%Y-%m-%d')
            usuarioDocente = matricula.curso_Materia.materia.docente.user.username
            if (fecha_formateada >= pRango1 and 
                fecha_formateada <= pRango2 and usuarioDocente == pUser):
                lista_observaciones = []
                contador = 0
                for observ in observaciones_asistencia:
                    contador += 1
                    lista_observaciones.append(f"Observación {contador}:{observ.observacionEst}")
                #Formateo de fechas
                hora_inicio = matricula.curso_Materia.materia.horario.hora_inicio
                hora_inicio_format = hora_inicio.strftime("%I:%M:%S %p")
                hora_fin = matricula.curso_Materia.materia.horario.hora_fin
                hora_fin_format = hora_fin.strftime("%I:%M:%S %p")
                dataReporte.append({
                    "id": dataAsistencia["id"],
                    "Tipo_asistencia": dataAsistencia["tipo_asistencia"],
                    "Descripcion": dataAsistencia["descripcion"],
                    "Hora_llegada": dataAsistencia["hora_llegada"],
                    "FechaCreacion": dataAsistencia["asistenciaEst_created_at"],
                    "Nombre_estudiante": estudiante.user.first_name + " " + estudiante.user.last_name,
                    "Cc_estudiante": cedulaEst,
                    "Curso_matriculado": curso.nombre_curso,
                    "Periodo": periodo.nombre_periodo,
                    "Materia": matricula.curso_Materia.materia.nombre_materia,
                    "Horario": f"{hora_inicio_format} a {hora_fin_format}",
                    "Docente": matricula.curso_Materia.materia.docente.user.first_name + " " + matricula.curso_Materia.materia.docente.user.last_name,
                    "Lista_Observaciones": lista_observaciones
                })
        return Response(dataReporte)
class AppApiReporteDiarioPDF(APIView):
    
    def get(self, request, format=None):
        asistencia = AsistenciaEstudiante.objects.all()
        observaciones = ObservacionesEstudiante.objects.all()
        serializer = AsistenciaEstudianteSerializer(asistencia, many=True)
        pUser = request.query_params.get("pUser", None)
        pRango1 = request.query_params.get("pRango1", None)
        pRango2 = request.query_params.get("pRango2", None)
        dataReporte = []

        if pRango1 is None or pRango2 is None or pUser is None:
            return Response(
                {"error": "Parametros incompletos(pRango1, pRango2, pUser)"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        
        for dataAsistencia in serializer.data:
            
            matricula = Matricula.objects.get(id=dataAsistencia["matricula_estudiante"])
            estudiante = Estudiante.objects.get(id=matricula.estudiante.id)
            cedulaEst = estudiante.estudiante_numero_Id
            curso = Curso.objects.get(id=matricula.curso_Materia.curso.id)
            periodo = Periodo.objects.get(id=matricula.curso_Materia.materia.periodo.id)
            fecha_objeto = datetime.strptime(dataAsistencia["asistenciaEst_created_at"], '%Y-%m-%dT%H:%M:%S.%fZ')
            observaciones_asistencia = observaciones.filter(asistenciaEst_id=dataAsistencia["id"])  # Filtrar observaciones por asistencia
            fecha_formateada = fecha_objeto.strftime('%Y-%m-%d')
            usuarioDocente = matricula.curso_Materia.materia.docente.user.username
            if (fecha_formateada >= pRango1 and 
                fecha_formateada <= pRango2 and usuarioDocente == pUser):
                lista_observaciones = []
                contador = 0
                for observ in observaciones_asistencia:
                    contador += 1
                    lista_observaciones.append(f"Observación {contador}:{observ.observacionEst}")
                #Formateo de fechas
                hora_inicio = matricula.curso_Materia.materia.horario.hora_inicio
                hora_inicio_format = hora_inicio.strftime("%I:%M:%S %p")
                hora_fin = matricula.curso_Materia.materia.horario.hora_fin
                hora_fin_format = hora_fin.strftime("%I:%M:%S %p")
                dataReporte.append({
                    "id": dataAsistencia["id"],
                    "Tipo_asistencia": dataAsistencia["tipo_asistencia"],
                    "Descripcion": dataAsistencia["descripcion"],
                    "Hora_llegada": dataAsistencia["hora_llegada"],
                    "FechaCreacion": dataAsistencia["asistenciaEst_created_at"],
                    "Nombre_estudiante": estudiante.user.first_name + " " + estudiante.user.last_name,
                    "Cc_estudiante": cedulaEst,
                    "Curso_matriculado": curso.nombre_curso,
                    "Periodo": periodo.nombre_periodo,
                    "Materia": matricula.curso_Materia.materia.nombre_materia,
                    "Horario": f"{hora_inicio_format} a {hora_fin_format}",
                    "Docente": matricula.curso_Materia.materia.docente.user.first_name + " " + matricula.curso_Materia.materia.docente.user.last_name,
                    "Lista_Observaciones": lista_observaciones
                })
        
        # Creamos un objeto BytesIO para almacenar el PDF generado
        buffer = BytesIO()
        # Creamos un objeto SimpleDocTemplate con el buffer como archivo de destino y tamaño de página oficio
        pdf = SimpleDocTemplate(buffer, pagesize=landscape(letter))

        # Lista de datos
        tabla_datos = []
        # Encabezado de la tabla
        encabezado = ["Tipo de asistencia", "Descripción", "Hora de llegada",
                       "Nombre del estudiante", "Cédula del estudiante", "Curso matriculado",
                       "Materia", "Horario", "Docente", "Lista de Observaciones"]

        tabla_datos.append(encabezado)
        for item in dataReporte:
            fila = [
                #item["id"],
                item["Tipo_asistencia"],
                item["Descripcion"],
                item["Hora_llegada"],
                #item["FechaCreacion"],
                item["Nombre_estudiante"],
                item["Cc_estudiante"],
                item["Curso_matriculado"],
                #item["Periodo"],
                item["Materia"],
                item["Horario"],
                item["Docente"],
                '\n'.join(item["Lista_Observaciones"])  # Convertimos la lista de observaciones a una cadena separada por saltos de línea
        ]
            tabla_datos.append(fila)
        tabla = Table(tabla_datos)

        # Aplicamos estilos a la tabla
        estilo = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),  # Color de fondo para el encabezado
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),  # Color de texto para el encabezado
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Alineación centrada para todas las celdas
            ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),  # Grosor y color de la línea interior de las celdas
            ('BOX', (0, 0), (-1, -1), 0.25, colors.black),  # Grosor y color del borde de las celdas
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Alineación vertical centrada
            ('WORDWRAP', (0, 0), (-1, -1), True),  # Ajuste de texto automático
        ])

        tabla.setStyle(estilo)

        # Agregamos la tabla al documento PDF
        elementos = []
        elementos.append(tabla)

        # Construimos el documento PDF
        pdf.build(elementos)

        # Volvemos al principio del buffer
        buffer.seek(0)

        # Creamos una respuesta HTTP con el contenido del buffer como archivo adjunto
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="archivo.pdf"'

        # Escribimos el contenido del buffer en la respuesta HTTP
        response.write(buffer.getvalue())

        return response

class AppApiReporteDiarioXLSX(APIView):
    
    
    def get(self, request, format = None):
        asistencia = AsistenciaEstudiante.objects.all()
        observaciones = ObservacionesEstudiante.objects.all()
        serializer = AsistenciaEstudianteSerializer(asistencia, many=True)
        pUser = request.query_params.get("pUser", None)
        pRango1 = request.query_params.get("pRango1", None)
        pRango2 = request.query_params.get("pRango2", None)
        dataReporte = []

        if pRango1 is None or pRango2 is None or pUser is None:
            return Response(
                {"error": "Parametros incompletos(pRango1, pRango2, pUser)"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        
        for dataAsistencia in serializer.data:
            
            matricula = Matricula.objects.get(id=dataAsistencia["matricula_estudiante"])
            estudiante = Estudiante.objects.get(id=matricula.estudiante.id)
            cedulaEst = estudiante.estudiante_numero_Id
            curso = Curso.objects.get(id=matricula.curso_Materia.curso.id)
            periodo = Periodo.objects.get(id=matricula.curso_Materia.materia.periodo.id)
            fecha_objeto = datetime.strptime(dataAsistencia["asistenciaEst_created_at"], '%Y-%m-%dT%H:%M:%S.%fZ')
            observaciones_asistencia = observaciones.filter(asistenciaEst_id=dataAsistencia["id"])  # Filtrar observaciones por asistencia
            fecha_formateada = fecha_objeto.strftime('%Y-%m-%d')
            usuarioDocente = matricula.curso_Materia.materia.docente.user.username
            if (fecha_formateada >= pRango1 and 
                fecha_formateada <= pRango2 and usuarioDocente == pUser):
                lista_observaciones = []
                contador = 0
                for observ in observaciones_asistencia:
                    contador += 1
                    lista_observaciones.append(f"Observación {contador}:{observ.observacionEst}")
                #Formateo de fechas
                hora_inicio = matricula.curso_Materia.materia.horario.hora_inicio
                hora_inicio_format = hora_inicio.strftime("%I:%M:%S %p")
                hora_fin = matricula.curso_Materia.materia.horario.hora_fin
                hora_fin_format = hora_fin.strftime("%I:%M:%S %p")
                dataReporte.append({
                    "id": dataAsistencia["id"],
                    "Tipo_asistencia": dataAsistencia["tipo_asistencia"],
                    "Descripcion": dataAsistencia["descripcion"],
                    "Hora_llegada": dataAsistencia["hora_llegada"],
                    "FechaCreacion": dataAsistencia["asistenciaEst_created_at"],
                    "Nombre_estudiante": estudiante.user.first_name + " " + estudiante.user.last_name,
                    "Cc_estudiante": cedulaEst,
                    "Curso_matriculado": curso.nombre_curso,
                    "Periodo": periodo.nombre_periodo,
                    "Materia": matricula.curso_Materia.materia.nombre_materia,
                    "Horario": f"{hora_inicio_format} a {hora_fin_format}",
                    "Docente": matricula.curso_Materia.materia.docente.user.first_name + " " + matricula.curso_Materia.materia.docente.user.last_name,
                    "Lista_Observaciones": lista_observaciones
                })        
        # Creamos un libro de trabajo de Excel
        wb = Workbook()
        ws = wb.active
        # Creamos el encabezado
        header = ["ID", "Tipo de asistencia", "Descripción", "Hora de llegada", "Fecha de creación",
                    "Nombre del estudiante", "Cédula del estudiante", "Curso matriculado", "Período",
                    "Materia", "Horario", "Docente"]

        # Agregamos el encabezado a la primera fila
        ws.append(header)
        # Añadimos los datos a las filas
        for item in dataReporte:
            # Creamos una fila para cada registro
            row = [
                item["id"],
                item["Tipo_asistencia"],
                item["Descripcion"],
                item["Hora_llegada"],
                item["FechaCreacion"],
                item["Nombre_estudiante"],
                item["Cc_estudiante"],
                item["Curso_matriculado"],
                item["Periodo"],
                item["Materia"],
                item["Horario"],
                item["Docente"]
            ]
            
            # Agregamos las observaciones en columnas adicionales
            for i, observacion in enumerate(item["Lista_Observaciones"]):
                row.append(observacion)

            # Añadimos la fila al libro de trabajo
            ws.append(row)

        # Creamos la respuesta HTTP
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="archivo.xlsx"'

        # Guardamos el libro de trabajo en la respuesta
        wb.save(response)

        return response


class AppApiReportePorCurso(APIView):
    def get(self, request, format=None):
        asistencia =  AsistenciaEstudiante.objects.all()
        observaciones = ObservacionesEstudiante.objects.all()
        serializer = AsistenciaEstudianteSerializer(asistencia, many=True)
        #Parametros para obtener el reporte filtrado.
        pUser = request.query_params.get("pUser", None)
        pMateria =  int(request.query_params.get("pMateria", None))
        pCurso = int(request.query_params.get("pCurso", None))
        pRango1 = request.query_params.get("pRango1", None)
        pRango2 = request.query_params.get("pRango2", None)
        

        dataReporte = []
        # Parametro de consulta
        if pMateria is None or pRango1 is None or pRango2 is None or pCurso is None or pUser is None:
            return Response(
                {"error": "Parametros incompletos(pCurso, pMateria, pRango1, pRango2)"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        # Recorridos
        for dataAsistencia in serializer.data:
            
            matricula = Matricula.objects.get(id=dataAsistencia["matricula_estudiante"])
            estudiante = Estudiante.objects.get(id=matricula.estudiante.id)
            cedulaEst =  estudiante.estudiante_numero_Id
            curso = Curso.objects.get(id=matricula.curso_Materia.curso.id)
            idCurso = matricula.curso_Materia.curso.id
            idMateria = matricula.curso_Materia.materia.id
            periodo = Periodo.objects.get(id=matricula.curso_Materia.materia.periodo.id)
            fecha_objeto = datetime.strptime(dataAsistencia["asistenciaEst_created_at"], '%Y-%m-%dT%H:%M:%S.%fZ')
            fecha_formateada = fecha_objeto.strftime('%Y-%m-%d')
            usuarioDocente = matricula.curso_Materia.materia.docente.user.username

            print(fecha_formateada >= pRango1)
            print(fecha_formateada <= pRango2 )
            print(pMateria == idMateria )
            print(idCurso == pCurso )
            print(usuarioDocente == pUser)
            
            if (fecha_formateada >= pRango1 and 
                fecha_formateada <= pRango2 and
                pMateria == idMateria and 
                idCurso == pCurso and
                usuarioDocente == pUser):
                
                observaciones = ObservacionesEstudiante.objects.all()
                lista_observaciones = []
                contador = 0
                for observ in observaciones:
                    if observ.asistenciaEst.id == dataAsistencia["id"]:
                        contador=contador + 1
                        lista_observaciones.append(f"Observación {contador}:{observ.observacionEst}")
                #Formateo de fechas
                hora_inicio = matricula.curso_Materia.materia.horario.hora_inicio
                hora_inicio_format = hora_inicio.strftime("%I:%M:%S %p")
                hora_fin = matricula.curso_Materia.materia.horario.hora_fin
                hora_fin_format = hora_fin.strftime("%I:%M:%S %p")
                print()
                dataReporte.append({
                    "id": dataAsistencia["id"],
                    "Tipo_asistencia": dataAsistencia["tipo_asistencia"],
                    "Descripcion": dataAsistencia["descripcion"],
                    "Hora_llegada": dataAsistencia["hora_llegada"],
                    "FechaCreacion": dataAsistencia["asistenciaEst_created_at"],
                    "Nombre_estudiante":estudiante.user.first_name + " " + estudiante.user.last_name,
                    "Cc_estudiante": cedulaEst,
                    "id_curso": matricula.curso_Materia.curso.id,
                    "Curso_matriculado": curso.nombre_curso,
                    "Periodo": periodo.nombre_periodo,
                    "id_Materia": matricula.curso_Materia.materia.id,
                    "Materia": matricula.curso_Materia.materia.nombre_materia,
                    "Horario": f"{hora_inicio_format} a {hora_fin_format}",
                    "Docente": matricula.curso_Materia.materia.docente.user.first_name + " " + matricula.curso_Materia.materia.docente.user.last_name,
                    "Lista_Observaciones": lista_observaciones
                })
        return Response(dataReporte)
    
class AppApiReportePorCursoPDF(APIView):
    def get(self, request, format=None):
        asistencia =  AsistenciaEstudiante.objects.all()
        observaciones = ObservacionesEstudiante.objects.all()
        serializer = AsistenciaEstudianteSerializer(asistencia, many=True)
        #Parametros para obtener el reporte filtrado.
        pUser = request.query_params.get("pUser", None)
        pMateria =  int(request.query_params.get("pMateria", None))
        pCurso = int(request.query_params.get("pCurso", None))
        pRango1 = request.query_params.get("pRango1", None)
        pRango2 = request.query_params.get("pRango2", None)
        

        dataReporte = []
        # Parametro de consulta
        if pMateria is None or pRango1 is None or pRango2 is None or pCurso is None or pUser is None:
            return Response(
                {"error": "Parametros incompletos(pCurso, pMateria, pRango1, pRango2)"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        # Recorridos
        for dataAsistencia in serializer.data:
            
            matricula = Matricula.objects.get(id=dataAsistencia["matricula_estudiante"])
            estudiante = Estudiante.objects.get(id=matricula.estudiante.id)
            cedulaEst =  estudiante.estudiante_numero_Id
            curso = Curso.objects.get(id=matricula.curso_Materia.curso.id)
            idCurso = matricula.curso_Materia.curso.id
            idMateria = matricula.curso_Materia.materia.id
            periodo = Periodo.objects.get(id=matricula.curso_Materia.materia.periodo.id)
            fecha_objeto = datetime.strptime(dataAsistencia["asistenciaEst_created_at"], '%Y-%m-%dT%H:%M:%S.%fZ')
            fecha_formateada = fecha_objeto.strftime('%Y-%m-%d')
            usuarioDocente = matricula.curso_Materia.materia.docente.user.username

            print(fecha_formateada >= pRango1)
            print(fecha_formateada <= pRango2 )
            print(pMateria == idMateria )
            print(idCurso == pCurso )
            print(usuarioDocente == pUser)
            
            if (fecha_formateada >= pRango1 and 
                fecha_formateada <= pRango2 and
                pMateria == idMateria and 
                idCurso == pCurso and
                usuarioDocente == pUser):
                
                observaciones = ObservacionesEstudiante.objects.all()
                lista_observaciones = []
                contador = 0
                for observ in observaciones:
                    if observ.asistenciaEst.id == dataAsistencia["id"]:
                        contador=contador + 1
                        lista_observaciones.append(f"Observación {contador}:{observ.observacionEst}")
                #Formateo de fechas
                hora_inicio = matricula.curso_Materia.materia.horario.hora_inicio
                hora_inicio_format = hora_inicio.strftime("%I:%M:%S %p")
                hora_fin = matricula.curso_Materia.materia.horario.hora_fin
                hora_fin_format = hora_fin.strftime("%I:%M:%S %p")
                print()
                dataReporte.append({
                    "id": dataAsistencia["id"],
                    "Tipo_asistencia": dataAsistencia["tipo_asistencia"],
                    "Descripcion": dataAsistencia["descripcion"],
                    "Hora_llegada": dataAsistencia["hora_llegada"],
                    "FechaCreacion": dataAsistencia["asistenciaEst_created_at"],
                    "Nombre_estudiante":estudiante.user.first_name + " " + estudiante.user.last_name,
                    "Cc_estudiante": cedulaEst,
                    "id_curso": matricula.curso_Materia.curso.id,
                    "Curso_matriculado": curso.nombre_curso,
                    "Periodo": periodo.nombre_periodo,
                    "id_Materia": matricula.curso_Materia.materia.id,
                    "Materia": matricula.curso_Materia.materia.nombre_materia,
                    "Horario": f"{hora_inicio_format} a {hora_fin_format}",
                    "Docente": matricula.curso_Materia.materia.docente.user.first_name + " " + matricula.curso_Materia.materia.docente.user.last_name,
                    "Lista_Observaciones": lista_observaciones
                })
        # Creamos un objeto BytesIO para almacenar el PDF generado
        buffer = BytesIO()
        # Creamos un objeto SimpleDocTemplate con el buffer como archivo de destino y tamaño de página oficio
        pdf = SimpleDocTemplate(buffer, pagesize=landscape(letter))

        # Lista de datos
        tabla_datos = []
        # Encabezado de la tabla
        encabezado = ["Tipo de asistencia", "Descripción", "Hora de llegada",
                       "Nombre del estudiante", "Cédula del estudiante", "Curso matriculado",
                       "Materia", "Horario", "Docente", "Lista de Observaciones"]

        tabla_datos.append(encabezado)
        for item in dataReporte:
            fila = [
                #item["id"],
                item["Tipo_asistencia"],
                item["Descripcion"],
                item["Hora_llegada"],
                #item["FechaCreacion"],
                item["Nombre_estudiante"],
                item["Cc_estudiante"],
                item["Curso_matriculado"],
                #item["Periodo"],
                item["Materia"],
                item["Horario"],
                item["Docente"],
                '\n'.join(item["Lista_Observaciones"])  # Convertimos la lista de observaciones a una cadena separada por saltos de línea
        ]
            tabla_datos.append(fila)
        tabla = Table(tabla_datos)

        # Aplicamos estilos a la tabla
        estilo = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),  # Color de fondo para el encabezado
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),  # Color de texto para el encabezado
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Alineación centrada para todas las celdas
            ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),  # Grosor y color de la línea interior de las celdas
            ('BOX', (0, 0), (-1, -1), 0.25, colors.black),  # Grosor y color del borde de las celdas
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Alineación vertical centrada
            ('WORDWRAP', (0, 0), (-1, -1), True),  # Ajuste de texto automático
        ])

        tabla.setStyle(estilo)

        # Agregamos la tabla al documento PDF
        elementos = []
        elementos.append(tabla)

        # Construimos el documento PDF
        pdf.build(elementos)

        # Volvemos al principio del buffer
        buffer.seek(0)

        # Creamos una respuesta HTTP con el contenido del buffer como archivo adjunto
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="archivo.pdf"'

        # Escribimos el contenido del buffer en la respuesta HTTP
        response.write(buffer.getvalue())

        return response
        


class AppApiReportePorCursoXLSX(APIView):
    def get(self, request, format=None):
        asistencia =  AsistenciaEstudiante.objects.all()
        observaciones = ObservacionesEstudiante.objects.all()
        serializer = AsistenciaEstudianteSerializer(asistencia, many=True)
        #Parametros para obtener el reporte filtrado.
        pUser = request.query_params.get("pUser", None)
        pMateria =  int(request.query_params.get("pMateria", None))
        pCurso = int(request.query_params.get("pCurso", None))
        pRango1 = request.query_params.get("pRango1", None)
        pRango2 = request.query_params.get("pRango2", None)
        

        dataReporte = []
        # Parametro de consulta
        if pMateria is None or pRango1 is None or pRango2 is None or pCurso is None or pUser is None:
            return Response(
                {"error": "Parametros incompletos(pCurso, pMateria, pRango1, pRango2, pUser)"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        # Recorridos
        for dataAsistencia in serializer.data:
            
            matricula = Matricula.objects.get(id=dataAsistencia["matricula_estudiante"])
            estudiante = Estudiante.objects.get(id=matricula.estudiante.id)
            cedulaEst =  estudiante.estudiante_numero_Id
            curso = Curso.objects.get(id=matricula.curso_Materia.curso.id)
            idCurso = matricula.curso_Materia.curso.id
            idMateria = matricula.curso_Materia.materia.id
            periodo = Periodo.objects.get(id=matricula.curso_Materia.materia.periodo.id)
            fecha_objeto = datetime.strptime(dataAsistencia["asistenciaEst_created_at"], '%Y-%m-%dT%H:%M:%S.%fZ')
            fecha_formateada = fecha_objeto.strftime('%Y-%m-%d')
            usuarioDocente = matricula.curso_Materia.materia.docente.user.username

            print(fecha_formateada >= pRango1)
            print(fecha_formateada <= pRango2 )
            print(pMateria == idMateria )
            print(idCurso == pCurso )
            print(usuarioDocente == pUser)
            
            if (fecha_formateada >= pRango1 and 
                fecha_formateada <= pRango2 and
                pMateria == idMateria and 
                idCurso == pCurso and
                usuarioDocente == pUser):
                
                observaciones = ObservacionesEstudiante.objects.all()
                lista_observaciones = []
                contador = 0
                for observ in observaciones:
                    if observ.asistenciaEst.id == dataAsistencia["id"]:
                        contador=contador + 1
                        lista_observaciones.append(f"Observación {contador}:{observ.observacionEst}")
                #Formateo de fechas
                hora_inicio = matricula.curso_Materia.materia.horario.hora_inicio
                hora_inicio_format = hora_inicio.strftime("%I:%M:%S %p")
                hora_fin = matricula.curso_Materia.materia.horario.hora_fin
                hora_fin_format = hora_fin.strftime("%I:%M:%S %p")
                print()
                dataReporte.append({
                    "id": dataAsistencia["id"],
                    "Tipo_asistencia": dataAsistencia["tipo_asistencia"],
                    "Descripcion": dataAsistencia["descripcion"],
                    "Hora_llegada": dataAsistencia["hora_llegada"],
                    "FechaCreacion": dataAsistencia["asistenciaEst_created_at"],
                    "Nombre_estudiante":estudiante.user.first_name + " " + estudiante.user.last_name,
                    "Cc_estudiante": cedulaEst,
                    "id_curso": matricula.curso_Materia.curso.id,
                    "Curso_matriculado": curso.nombre_curso,
                    "Periodo": periodo.nombre_periodo,
                    "id_Materia": matricula.curso_Materia.materia.id,
                    "Materia": matricula.curso_Materia.materia.nombre_materia,
                    "Horario": f"{hora_inicio_format} a {hora_fin_format}",
                    "Docente": matricula.curso_Materia.materia.docente.user.first_name + " " + matricula.curso_Materia.materia.docente.user.last_name,
                    "Lista_Observaciones": lista_observaciones
                })
# Creamos un libro de trabajo de Excel
        wb = Workbook()
        ws = wb.active
        # Creamos el encabezado
        header = ["ID", "Tipo de asistencia", "Descripción", "Hora de llegada", "Fecha de creación",
                    "Nombre del estudiante", "Cédula del estudiante", "Curso matriculado", "Período",
                    "Materia", "Horario", "Docente"]

        # Agregamos el encabezado a la primera fila
        ws.append(header)
        # Añadimos los datos a las filas
        for item in dataReporte:
            # Creamos una fila para cada registro
            row = [
                item["id"],
                item["Tipo_asistencia"],
                item["Descripcion"],
                item["Hora_llegada"],
                item["FechaCreacion"],
                item["Nombre_estudiante"],
                item["Cc_estudiante"],
                item["Curso_matriculado"],
                item["Periodo"],
                item["Materia"],
                item["Horario"],
                item["Docente"]
            ]
            
            # Agregamos las observaciones en columnas adicionales
            for i, observacion in enumerate(item["Lista_Observaciones"]):
                row.append(observacion)

            # Añadimos la fila al libro de trabajo
            ws.append(row)

        # Creamos la respuesta HTTP
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="archivo.xlsx"'

        # Guardamos el libro de trabajo en la respuesta
        wb.save(response)

        return response